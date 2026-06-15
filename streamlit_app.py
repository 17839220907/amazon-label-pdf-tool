import io
import json
import mimetypes
import os
import re
import shutil
import tempfile
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
import zipfile
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from string import Formatter

import fitz
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


FNSKU_REGEX = r"\bX[A-Z0-9]{9}\b"
FILENAME_TEMPLATE = "{SKU}-{FNSKU}.pdf"
LABEL_TEXT_LINE2_TEMPLATE = "{款式}"

PDF_TEXT_COVER_TOP_RATIO = 0.535
PDF_FNSKU_BASELINE_RATIO = 0.660
PDF_LINE1_BASELINE_RATIO = 0.783
PDF_LINE2_BASELINE_RATIO = 0.895
PDF_SIDE_MARGIN = 4

CJK_FONT_CANDIDATES = []
CJK_BOLD_STROKE_WIDTH = 0.05

FEISHU_INDEX_RANGE_DEFAULT = "A1:G50000"
SKIP_SHEET_NAMES = {"更新日志", "更新记录", "日志", "说明", "README"}
PROCESS_HISTORY_DIR = Path(os.environ.get("LABEL_TOOL_HISTORY_DIR", "/tmp/amazon_label_pdf_history"))
PROCESS_HISTORY_LIMIT = 3
JOB_DIR = Path(os.environ.get("LABEL_TOOL_JOB_DIR", "/tmp/amazon_label_pdf_jobs"))
JOB_HISTORY_LIMIT = 10
UPLOAD_WORKERS = max(1, int(os.environ.get("LABEL_TOOL_UPLOAD_WORKERS", "5")))
FEISHU_REQUEST_TIMEOUT = max(5, int(os.environ.get("LABEL_TOOL_FEISHU_TIMEOUT", "45")))

LOCAL_ILLEGAL_FILENAME_CHARS = r'\/:*?"<>|'

LOG_HEADERS = [
    "处理时间",
    "原文件名",
    "页码",
    "识别FNSKU",
    "匹配状态",
    "失败原因",
    "标签第一行",
    "标签第二行",
    "飞书正式文件名",
    "本地备份文件名",
    "飞书文件token",
    "处理动作",
]

REQUIRED_INDEX_HEADERS = ["FNSKU", "SKU", "款式", "品牌", "型号", "产品类型"]
REQUIRED_INFO_FIELDS = ["SKU", "款式", "品牌", "型号", "产品类型"]
OPTIONAL_INFO_FIELDS = ["颜色"]
SKIPPED_STATUS = "已跳过"
SUCCESS_STATUS = "成功"


class FatalError(Exception):
    pass


class FileNameError(Exception):
    pass


class FeishuError(Exception):
    pass


class JobCancelled(Exception):
    pass


def cell_to_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_fnsku(value):
    return cell_to_text(value).upper()


def is_valid_fnsku(value):
    return re.fullmatch(FNSKU_REGEX, value, flags=re.IGNORECASE) is not None


def contains_cjk(text):
    return any("\u4e00" <= char <= "\u9fff" for char in text)


def is_cjk_char(char):
    return "\u4e00" <= char <= "\u9fff"


def find_cjk_fontfile():
    for font_path in CJK_FONT_CANDIDATES:
        if Path(font_path).exists():
            return font_path
    return ""


def ensure_pdf_suffix(filename):
    if filename.lower().endswith(".pdf"):
        return filename[:-4] + ".pdf"
    return filename + ".pdf"


def normalize_display_filename(filename):
    filename = cell_to_text(filename)
    filename = re.sub(r"[\r\n\t]+", " ", filename)
    filename = re.sub(r"\s+", "-", filename)
    filename = filename.strip(" .-")
    if not filename:
        raise FileNameError("文件名为空")

    filename = ensure_pdf_suffix(filename)
    if filename.lower() == ".pdf":
        raise FileNameError("文件名为空")
    return filename


def make_safe_local_filename(display_filename, fnsku):
    filename = normalize_display_filename(display_filename)
    for char in LOCAL_ILLEGAL_FILENAME_CHARS:
        filename = filename.replace(char, "-")
    filename = re.sub(r"-+", "-", filename)
    filename = filename.strip(" .-")
    if not filename:
        filename = f"{fnsku}.pdf"
    return ensure_pdf_suffix(filename)


def get_template_fields(template):
    fields = []
    for _, field_name, _, _ in Formatter().parse(template):
        if not field_name:
            continue
        simple_name = field_name.split(".", 1)[0].split("[", 1)[0]
        if simple_name and simple_name not in fields:
            fields.append(simple_name)
    return fields


def build_display_filename(row, fnsku):
    values = {}
    missing_fields = []

    for field in get_template_fields(FILENAME_TEMPLATE):
        if field == "FNSKU":
            value = fnsku
        else:
            value = cell_to_text(row.get(field))

        if not value:
            missing_fields.append(field)
        values[field] = value

    if missing_fields:
        raise FileNameError("生成文件名所需信息缺失：" + "、".join(missing_fields))

    try:
        filename = normalize_display_filename(FILENAME_TEMPLATE.format(**values))
    except KeyError as exc:
        raise FileNameError(f"文件名模板字段不存在：{exc}") from exc

    if fnsku.upper() not in filename.upper():
        raise FileNameError("输出文件名未包含识别到的 FNSKU")
    return filename


def build_label_lines(row):
    missing_fields = [field for field in REQUIRED_INFO_FIELDS if not cell_to_text(row.get(field))]
    if missing_fields:
        raise FileNameError("标签信息缺失：" + "、".join(missing_fields))

    brand = cell_to_text(row.get("品牌"))
    model = cell_to_text(row.get("型号"))
    product_type = cell_to_text(row.get("产品类型"))
    color = cell_to_text(row.get("颜色"))

    line1 = f"{brand} for {model} {product_type}"
    if color:
        line1 = f"{line1}, {color}"

    try:
        line2 = LABEL_TEXT_LINE2_TEMPLATE.format(**row)
    except KeyError as exc:
        raise FileNameError(f"标签文字模板字段不存在：{exc}") from exc

    line1 = cell_to_text(line1)
    line2 = cell_to_text(line2)
    if not line1 or not line2:
        raise FileNameError("标签文字为空")
    return line1, line2


def get_draw_font(text, cjk_fontfile):
    if contains_cjk(text):
        if cjk_fontfile:
            return "labelcjk", cjk_fontfile, fitz.Font(fontfile=cjk_fontfile)
        return "china-ss", "", fitz.Font(fontname="china-ss")
    return "helv", "", fitz.Font(fontname="helv")


def split_font_runs(text):
    runs = []
    current_text = ""
    current_is_cjk = None

    for char in text:
        char_is_cjk = is_cjk_char(char)
        if current_text and char_is_cjk != current_is_cjk:
            runs.append((current_text, current_is_cjk))
            current_text = char
        else:
            current_text += char
        current_is_cjk = char_is_cjk

    if current_text:
        runs.append((current_text, current_is_cjk))
    return runs


def measure_runs_width(runs, font_size, cjk_fontfile):
    total_width = 0
    prepared_runs = []

    for run_text, _ in runs:
        fontname, fontfile, font = get_draw_font(run_text, cjk_fontfile)
        width = font.text_length(run_text, fontsize=font_size)
        prepared_runs.append((run_text, fontname, fontfile, width))
        total_width += width

    return total_width, prepared_runs


def fit_runs_font_size(runs, cjk_fontfile, max_width, start_size, min_size):
    font_size = start_size
    while font_size > min_size:
        total_width, _ = measure_runs_width(runs, font_size, cjk_fontfile)
        if total_width <= max_width:
            break
        font_size -= 0.2
    return max(font_size, min_size)


def draw_centered_text(page, text, baseline_y, start_size, min_size, cjk_fontfile):
    text = cell_to_text(text)
    if not text:
        return

    runs = split_font_runs(text)
    max_width = page.rect.width - PDF_SIDE_MARGIN * 2 - 2
    font_size = fit_runs_font_size(runs, cjk_fontfile, max_width, start_size, min_size)
    text_width, prepared_runs = measure_runs_width(runs, font_size, cjk_fontfile)
    x = (page.rect.width - text_width) / 2

    for run_text, fontname, fontfile, width in prepared_runs:
        kwargs = {
            "fontsize": font_size,
            "fontname": fontname,
            "color": (0, 0, 0),
            "fill": (0, 0, 0),
            "overlay": True,
        }
        if contains_cjk(run_text) and CJK_BOLD_STROKE_WIDTH > 0:
            kwargs["render_mode"] = 2
            kwargs["border_width"] = CJK_BOLD_STROKE_WIDTH
        if fontfile:
            kwargs["fontfile"] = fontfile

        page.insert_text(fitz.Point(x, baseline_y), run_text, **kwargs)
        x += width


def rewrite_label_pdf(source_path, target_path, page_index, fnsku, label_line1, label_line2):
    target_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = target_path.parent / f".{target_path.stem}.{uuid.uuid4().hex}.tmp.pdf"
    cjk_fontfile = find_cjk_fontfile()

    try:
        with fitz.open(source_path) as source_document:
            if source_document.page_count == 0:
                raise RuntimeError("PDF 没有页面")
            if page_index < 0 or page_index >= source_document.page_count:
                raise RuntimeError(f"PDF 页码不存在：第 {page_index + 1} 页")

            document = fitz.open()
            try:
                document.insert_pdf(source_document, from_page=page_index, to_page=page_index)
                page = document[0]
                rect = page.rect
                scale = rect.width / 142.08 if rect.width else 1

                erase_rect = fitz.Rect(
                    PDF_SIDE_MARGIN,
                    rect.height * PDF_TEXT_COVER_TOP_RATIO,
                    rect.width - PDF_SIDE_MARGIN,
                    rect.height - 2,
                )
                page.add_redact_annot(erase_rect, fill=(1, 1, 1))
                page.apply_redactions()

                draw_centered_text(
                    page,
                    fnsku,
                    rect.height * PDF_FNSKU_BASELINE_RATIO,
                    10.5 * scale,
                    7.5 * scale,
                    cjk_fontfile,
                )
                draw_centered_text(
                    page,
                    label_line1,
                    rect.height * PDF_LINE1_BASELINE_RATIO,
                    6.9 * scale,
                    4.8 * scale,
                    cjk_fontfile,
                )
                draw_centered_text(
                    page,
                    label_line2,
                    rect.height * PDF_LINE2_BASELINE_RATIO,
                    7.4 * scale,
                    5.0 * scale,
                    cjk_fontfile,
                )

                document.save(temp_path, garbage=4, deflate=True)
            finally:
                document.close()

        os.replace(temp_path, target_path)
    except Exception:
        if temp_path.exists():
            temp_path.unlink()
        raise

    return "生成修改后PDF"


def format_sheet_row(sheet_name, row_number):
    return f"「{sheet_name}」第 {row_number} 行"


def sheet_should_skip(sheet_name):
    return cell_to_text(sheet_name) in SKIP_SHEET_NAMES


def build_header_to_index(values):
    if not values:
        return {}
    header_values = [cell_to_text(value) for value in values[0]]
    return {header: idx for idx, header in enumerate(header_values) if header}


def sheet_looks_like_index(header_to_index):
    return any(header in header_to_index for header in REQUIRED_INDEX_HEADERS)


def parse_index_sheets(sheet_values_list):
    missing_fnsku_rows = []
    invalid_fnsku_rows = []
    skipped_same_duplicate_rows = []
    skipped_sheet_names = []
    used_sheet_names = []
    index_rows = {}
    duplicate_conflicts = defaultdict(list)
    duplicate_check_fields = REQUIRED_INFO_FIELDS + OPTIONAL_INFO_FIELDS

    for sheet_name, values in sheet_values_list:
        sheet_name = cell_to_text(sheet_name) or "未命名Sheet"
        if sheet_should_skip(sheet_name):
            skipped_sheet_names.append(sheet_name)
            continue

        header_to_index = build_header_to_index(values)
        if not sheet_looks_like_index(header_to_index):
            skipped_sheet_names.append(sheet_name)
            continue

        missing_headers = [header for header in REQUIRED_INDEX_HEADERS if header not in header_to_index]
        if missing_headers:
            raise FatalError(f"工作表「{sheet_name}」像索引表，但缺少表头：" + "、".join(missing_headers))

        used_sheet_names.append(sheet_name)

        for row_number, row_values in enumerate(values[1:], start=2):
            row = {}
            has_any_value = False

            for header, idx in header_to_index.items():
                value = row_values[idx] if idx < len(row_values) else None
                text = cell_to_text(value)
                row[header] = text
                if text:
                    has_any_value = True

            if not has_any_value:
                continue

            row_ref = format_sheet_row(sheet_name, row_number)
            fnsku = normalize_fnsku(row.get("FNSKU"))
            if not fnsku:
                missing_fnsku_rows.append(row_ref)
                continue

            if not is_valid_fnsku(fnsku):
                invalid_fnsku_rows.append((row_ref, fnsku))
                continue

            row["FNSKU"] = fnsku
            row["_索引位置"] = row_ref

            existing_row = index_rows.get(fnsku)
            if existing_row:
                is_same_content = all(
                    cell_to_text(existing_row.get(field)) == cell_to_text(row.get(field))
                    for field in duplicate_check_fields
                )
                if is_same_content:
                    skipped_same_duplicate_rows.append((row_ref, fnsku, existing_row["_索引位置"]))
                    continue

                if not duplicate_conflicts[fnsku]:
                    duplicate_conflicts[fnsku].append(existing_row["_索引位置"])
                duplicate_conflicts[fnsku].append(row_ref)
                continue

            index_rows[fnsku] = row

    if not used_sheet_names:
        raise FatalError("没有找到索引 Sheet：请确认至少一个 Sheet 第一行包含 FNSKU、SKU 等表头")

    warnings = []
    if missing_fnsku_rows:
        sample_text = "；".join(missing_fnsku_rows[:10])
        if len(missing_fnsku_rows) > 10:
            sample_text += "；..."
        warnings.append(f"已忽略 {len(missing_fnsku_rows)} 行缺少 FNSKU 的半填写索引行（{sample_text}）")

    if invalid_fnsku_rows:
        sample_text = "；".join(f"{row_ref} {fnsku}" for row_ref, fnsku in invalid_fnsku_rows[:10])
        if len(invalid_fnsku_rows) > 10:
            sample_text += "；..."
        warnings.append(f"已忽略 {len(invalid_fnsku_rows)} 行无效 FNSKU（{sample_text}）")

    if skipped_same_duplicate_rows:
        sample_text = "；".join(
            f"{fnsku} {row_ref}（与 {first_row_ref} 一致）"
            for row_ref, fnsku, first_row_ref in skipped_same_duplicate_rows[:10]
        )
        if len(skipped_same_duplicate_rows) > 10:
            sample_text += "；..."
        warnings.append(f"已忽略 {len(skipped_same_duplicate_rows)} 行完全相同的重复 FNSKU（{sample_text}）")

    if duplicate_conflicts:
        lines = ["索引表中 FNSKU 重复且信息不同，程序已停止："]
        for fnsku, row_refs in sorted(duplicate_conflicts.items()):
            lines.append(f"- {fnsku}：{', '.join(row_refs)}")
        raise FatalError("\n".join(lines))

    return {
        "rows": index_rows,
        "warnings": warnings,
        "used_sheets": used_sheet_names,
        "skipped_sheets": skipped_sheet_names,
    }


def urlopen_json(request):
    try:
        with urllib.request.urlopen(request, timeout=FEISHU_REQUEST_TIMEOUT) as response:
            raw = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise FeishuError(f"飞书接口 HTTP {exc.code}：{detail}") from exc
    except urllib.error.URLError as exc:
        raise FeishuError(f"飞书接口连接失败：{exc}") from exc

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise FeishuError(f"飞书接口返回不是 JSON：{raw[:300]}") from exc

    if data.get("code") not in (0, None):
        raise FeishuError(f"飞书接口返回错误：{data}")
    return data


def feishu_json_request(method, path, token=None, payload=None, query=None):
    url = "https://open.feishu.cn" + path
    if query:
        url += "?" + urllib.parse.urlencode(query)

    headers = {
        "User-Agent": "amazon-label-pdf-tool",
        "Accept": "application/json; charset=utf-8",
    }
    body = None
    if token:
        headers["Authorization"] = f"Bearer {token}"
    if payload is not None:
        headers["Content-Type"] = "application/json; charset=utf-8"
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")

    request = urllib.request.Request(url, data=body, headers=headers, method=method)
    return urlopen_json(request)


def get_secret_value(section, key, default=""):
    try:
        value = st.secrets.get(section, {}).get(key, default)
    except Exception:
        value = default
    return cell_to_text(value)


def load_feishu_config():
    config = {
        "app_id": get_secret_value("feishu", "app_id"),
        "app_secret": get_secret_value("feishu", "app_secret"),
        "spreadsheet_token": get_secret_value("feishu", "spreadsheet_token"),
        "output_folder_token": get_secret_value("feishu", "output_folder_token"),
        "index_range": get_secret_value("feishu", "index_range", FEISHU_INDEX_RANGE_DEFAULT),
    }
    missing = [key for key, value in config.items() if key != "index_range" and not value]
    if missing:
        raise FatalError("Streamlit Secrets 缺少飞书配置：" + "、".join(missing))
    return config


def get_feishu_tenant_access_token(config):
    data = feishu_json_request(
        "POST",
        "/open-apis/auth/v3/tenant_access_token/internal",
        payload={"app_id": config["app_id"], "app_secret": config["app_secret"]},
    )
    token = data.get("tenant_access_token")
    if not token:
        raise FeishuError(f"未获取到 tenant_access_token：{data}")
    return token


def get_nested_value(data, *keys):
    current = data
    for key in keys:
        if not isinstance(current, dict):
            return ""
        current = current.get(key)
    return current


def normalize_feishu_sheet_infos(data):
    payload = data.get("data", {})
    raw_sheets = (
        payload.get("sheets")
        or payload.get("items")
        or payload.get("sheet")
        or payload.get("spreadsheet", {}).get("sheets")
        or []
    )
    if isinstance(raw_sheets, dict):
        raw_sheets = list(raw_sheets.values())

    sheet_infos = []
    for item in raw_sheets:
        if not isinstance(item, dict):
            continue
        title = cell_to_text(
            item.get("title")
            or item.get("name")
            or item.get("sheet_name")
            or item.get("sheetName")
            or get_nested_value(item, "properties", "title")
        )
        sheet_id = cell_to_text(
            item.get("sheet_id")
            or item.get("sheetId")
            or item.get("id")
            or item.get("sheet_token")
            or get_nested_value(item, "properties", "sheet_id")
            or get_nested_value(item, "properties", "sheetId")
        )
        if title or sheet_id:
            sheet_infos.append({"title": title or sheet_id, "sheet_id": sheet_id or title})
    return sheet_infos


def get_feishu_sheet_infos(token, spreadsheet_token):
    errors = []
    for path in (
        f"/open-apis/sheets/v3/spreadsheets/{spreadsheet_token}/sheets/query",
        f"/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/metainfo",
    ):
        try:
            data = feishu_json_request("GET", path, token=token)
            sheet_infos = normalize_feishu_sheet_infos(data)
            if sheet_infos:
                return sheet_infos
        except FeishuError as exc:
            errors.append(str(exc))

    detail = "；".join(errors) if errors else "飞书没有返回工作表列表"
    raise FeishuError(f"读取飞书 Sheet 列表失败：{detail}")


def read_feishu_range_values(token, spreadsheet_token, range_text):
    encoded_range = urllib.parse.quote(range_text, safe="!:$")
    data = feishu_json_request(
        "GET",
        f"/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values/{encoded_range}",
        token=token,
    )
    value_range = data.get("data", {}).get("valueRange", {})
    return value_range.get("values") or []


def read_feishu_sheet_index(token, config):
    sheet_values_list = []
    for sheet_info in get_feishu_sheet_infos(token, config["spreadsheet_token"]):
        sheet_title = sheet_info["title"]
        if sheet_should_skip(sheet_title):
            sheet_values_list.append((sheet_title, []))
            continue

        range_text = f"{sheet_info['sheet_id']}!{config['index_range']}"
        values = read_feishu_range_values(token, config["spreadsheet_token"], range_text)
        sheet_values_list.append((sheet_title, values))

    return parse_index_sheets(sheet_values_list)


def make_multipart_form(fields, file_field, file_path, display_filename, upload_filename=None):
    boundary = "----label-feishu-" + uuid.uuid4().hex
    body = bytearray()

    for name, value in fields.items():
        body.extend(f"--{boundary}\r\n".encode())
        body.extend(f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode())
        body.extend(str(value).encode("utf-8"))
        body.extend(b"\r\n")

    upload_filename = upload_filename or display_filename
    content_type = mimetypes.guess_type(display_filename)[0] or "application/pdf"
    body.extend(f"--{boundary}\r\n".encode())
    body.extend(
        (
            f'Content-Disposition: form-data; name="{file_field}"; '
            f'filename="{upload_filename}"\r\n'
        ).encode("utf-8")
    )
    body.extend(f"Content-Type: {content_type}\r\n\r\n".encode())
    body.extend(file_path.read_bytes())
    body.extend(b"\r\n")
    body.extend(f"--{boundary}--\r\n".encode())

    return bytes(body), f"multipart/form-data; boundary={boundary}"


def feishu_upload_file(token, config, local_path, display_filename):
    size = local_path.stat().st_size
    upload_filename = make_safe_local_filename(display_filename, "upload")
    fields = {
        "file_name": display_filename,
        "parent_type": "explorer",
        "parent_node": config["output_folder_token"],
        "size": str(size),
    }
    body, content_type = make_multipart_form(
        fields,
        "file",
        local_path,
        display_filename,
        upload_filename=upload_filename,
    )
    request = urllib.request.Request(
        "https://open.feishu.cn/open-apis/drive/v1/medias/upload_all",
        data=body,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": content_type,
            "User-Agent": "amazon-label-pdf-tool",
        },
        method="POST",
    )
    data = urlopen_json(request)
    file_token = data.get("data", {}).get("file_token") or data.get("data", {}).get("token")
    if not file_token:
        raise FeishuError(f"飞书上传成功但未返回 file_token：{data}")
    return file_token


def feishu_list_folder_files(token, config):
    files = []
    page_token = ""
    while True:
        query = {"folder_token": config["output_folder_token"], "page_size": 200}
        if page_token:
            query["page_token"] = page_token

        data = feishu_json_request("GET", "/open-apis/drive/v1/files", token=token, query=query)
        payload = data.get("data", {})
        batch = payload.get("files") or payload.get("items") or []
        files.extend(batch)

        if not payload.get("has_more"):
            break
        page_token = payload.get("next_page_token") or payload.get("page_token") or ""
        if not page_token:
            break
    return files


def feishu_delete_file(token, file_token, file_type="file"):
    if not file_token:
        return
    feishu_json_request(
        "DELETE",
        f"/open-apis/drive/v1/files/{urllib.parse.quote(file_token)}",
        token=token,
        query={"type": file_type},
    )


def build_remote_fnsku_map(remote_files):
    remote_by_fnsku = defaultdict(list)
    pattern = re.compile(FNSKU_REGEX, re.IGNORECASE)
    for item in remote_files:
        name = cell_to_text(item.get("name") or item.get("file_name") or item.get("title"))
        token = cell_to_text(item.get("token") or item.get("file_token"))
        file_type = cell_to_text(item.get("type") or "file")
        timestamp = cell_to_text(
            item.get("modified_time")
            or item.get("updated_time")
            or item.get("created_time")
            or item.get("create_time")
        )
        if not name or not token:
            continue
        for match in pattern.finditer(name):
            remote_by_fnsku[match.group(0).upper()].append(
                {"name": name, "token": token, "type": file_type, "timestamp": timestamp}
            )
    return remote_by_fnsku


def cleanup_remote_duplicates(token, remote_by_fnsku):
    deleted = []
    for fnsku, files in list(remote_by_fnsku.items()):
        unique_files = []
        seen_tokens = set()
        for file_info in files:
            file_token = cell_to_text(file_info.get("token"))
            if not file_token or file_token in seen_tokens:
                continue
            seen_tokens.add(file_token)
            unique_files.append(file_info)

        if len(unique_files) <= 1:
            remote_by_fnsku[fnsku] = unique_files
            continue

        unique_files.sort(
            key=lambda item: (cell_to_text(item.get("timestamp")), cell_to_text(item.get("name"))),
            reverse=True,
        )
        keep_file = unique_files[0]
        for duplicate_file in unique_files[1:]:
            feishu_delete_file(token, duplicate_file["token"], duplicate_file.get("type") or "file")
            deleted.append(
                {
                    "fnsku": fnsku,
                    "deleted_name": duplicate_file.get("name", ""),
                    "kept_name": keep_file.get("name", ""),
                }
            )
        remote_by_fnsku[fnsku] = [keep_file]
    return deleted


def extract_fnsku_list(text, pattern):
    found = [match.group(0).upper() for match in pattern.finditer(text)]
    return sorted(set(found))


def natural_sort_key(name):
    parts = re.split(r"(\d+)", name.lower())
    return [int(part) if part.isdigit() else part for part in parts]


def safe_upload_name(filename, fallback_suffix):
    name = Path(filename).name.strip() or f"uploaded{fallback_suffix}"
    name = name.replace("\\", "_").replace("/", "_")
    return name


def unique_path(folder, filename):
    candidate = folder / filename
    if not candidate.exists():
        return candidate

    stem = candidate.stem
    suffix = candidate.suffix
    counter = 1
    while True:
        candidate = folder / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def make_log_row(
    source_name,
    page_number,
    fnsku,
    status,
    reason,
    display_filename,
    local_filename,
    feishu_file_token,
    action,
    label_line1="",
    label_line2="",
):
    return {
        "处理时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "原文件名": source_name,
        "页码": page_number,
        "识别FNSKU": fnsku,
        "匹配状态": status,
        "失败原因": reason,
        "标签第一行": label_line1,
        "标签第二行": label_line2,
        "飞书正式文件名": display_filename,
        "本地备份文件名": local_filename,
        "飞书文件token": feishu_file_token,
        "处理动作": action,
    }


def fail_row(
    source_name,
    category,
    fnsku,
    reason,
    page_number="",
    display_filename="",
    local_filename="",
    label_line1="",
    label_line2="",
):
    return make_log_row(
        source_name,
        page_number,
        fnsku,
        "失败",
        reason,
        display_filename,
        local_filename,
        "",
        f"异常({category}，未上传)",
        label_line1,
        label_line2,
    )


def skip_row(
    source_name,
    fnsku,
    reason,
    page_number="",
    display_filename="",
    local_filename="",
    feishu_file_token="",
    label_line1="",
    label_line2="",
):
    return make_log_row(
        source_name,
        page_number,
        fnsku,
        SKIPPED_STATUS,
        reason,
        display_filename,
        local_filename,
        feishu_file_token,
        f"跳过({reason})",
        label_line1,
        label_line2,
    )


def process_pdf_page(
    pdf_path,
    source_name,
    page_index,
    page_count,
    page_text,
    index_rows,
    fnsku_pattern,
    output_dir,
    batch_fnskus,
    feishu_token,
    feishu_config,
    remote_by_fnsku,
    delete_existing_same_fnsku,
    stage_callback=None,
    defer_upload=False,
):
    recognized_fnsku = ""
    page_number = page_index + 1

    def report_stage(message):
        if stage_callback:
            stage_callback(message)

    try:
        report_stage("正在提取 PDF 文字")
        text = cell_to_text(page_text)
        if not text:
            raise RuntimeError("无法提取 PDF 文字")

        report_stage("正在识别 FNSKU")
        fnsku_list = extract_fnsku_list(text, fnsku_pattern)
        if not fnsku_list:
            return fail_row(source_name, "未识别FNSKU", "", "未识别 FNSKU", page_number=page_number)

        if len(fnsku_list) > 1:
            recognized_fnsku = "、".join(fnsku_list)
            return fail_row(source_name, "多个FNSKU", recognized_fnsku, "识别到多个不同 FNSKU", page_number=page_number)

        recognized_fnsku = fnsku_list[0]
        report_stage(f"已识别 {recognized_fnsku}，正在匹配索引表")
        if recognized_fnsku in batch_fnskus:
            return skip_row(
                source_name,
                recognized_fnsku,
                "本次上传已处理过相同 FNSKU，已跳过，避免重复标签",
                page_number=page_number,
            )

        index_row = index_rows.get(recognized_fnsku)
        if not index_row:
            return fail_row(source_name, "索引无匹配", recognized_fnsku, "索引表无匹配", page_number=page_number)

        try:
            label_line1, label_line2 = build_label_lines(index_row)
            display_filename = build_display_filename(index_row, recognized_fnsku)
            local_filename = make_safe_local_filename(display_filename, recognized_fnsku)
        except FileNameError as exc:
            return fail_row(source_name, "信息异常", recognized_fnsku, str(exc), page_number=page_number)

        old_files = remote_by_fnsku.get(recognized_fnsku, [])
        if old_files:
            report_stage(f"{recognized_fnsku} 飞书已存在，跳过上传")
            old_file = old_files[0]
            return skip_row(
                source_name,
                recognized_fnsku,
                "飞书目标文件夹已存在同 FNSKU，未重复上传",
                page_number=page_number,
                display_filename=display_filename,
                local_filename="",
                feishu_file_token=old_file.get("token", ""),
                label_line1=label_line1,
                label_line2=label_line2,
            )

        report_stage(f"{recognized_fnsku} 正在生成新标签 PDF")
        local_path = unique_path(output_dir, local_filename)
        action = rewrite_label_pdf(
            pdf_path,
            local_path,
            page_index,
            recognized_fnsku,
            label_line1,
            label_line2,
        )

        batch_fnskus.add(recognized_fnsku)

        page_info = f"第 {page_number}/{page_count} 页" if page_count > 1 else str(page_number)
        row = make_log_row(
            source_name,
            page_info,
            recognized_fnsku,
            SUCCESS_STATUS,
            "",
            display_filename,
            local_path.name,
            "",
            f"{action}；等待上传飞书" if defer_upload else action,
            label_line1,
            label_line2,
        )
        row["_local_path"] = str(local_path)
        row["_needs_upload"] = True

        if defer_upload:
            return row

        report_stage(f"{recognized_fnsku} 正在上传到飞书")
        feishu_file_token = feishu_upload_file(feishu_token, feishu_config, local_path, display_filename)
        row["飞书文件token"] = feishu_file_token
        row["处理动作"] = f"{action}；上传飞书"
        remote_by_fnsku[recognized_fnsku] = [{"name": display_filename, "token": feishu_file_token, "type": "file"}]
        row["_needs_upload"] = False
        return row

    except RuntimeError as exc:
        return fail_row(source_name, "无法提取文字", recognized_fnsku, str(exc), page_number=page_number)
    except Exception as exc:
        return fail_row(source_name, "处理失败", recognized_fnsku, f"处理失败：{exc}", page_number=page_number)


def process_pdf_file(
    pdf_path,
    source_name,
    index_rows,
    fnsku_pattern,
    output_dir,
    batch_fnskus,
    feishu_token,
    feishu_config,
    remote_by_fnsku,
    delete_existing_same_fnsku,
    stage_callback=None,
    page_finished_callback=None,
    defer_upload=False,
):
    try:
        if stage_callback:
            stage_callback(source_name, 0, 0, "正在打开 PDF 并读取页数")
        with fitz.open(pdf_path) as document:
            page_count = document.page_count
            if page_count == 0:
                row = fail_row(source_name, "无法提取文字", "", "PDF 没有页面")
                if page_finished_callback:
                    page_finished_callback(source_name, 0, 1, row)
                return [row]
            page_texts = [document[page_index].get_text("text") or "" for page_index in range(page_count)]

        log_rows = []
        for page_index, page_text in enumerate(page_texts):
            def page_stage(message, page_index=page_index):
                if stage_callback:
                    stage_callback(source_name, page_index, page_count, message)

            page_stage("准备处理这一页")
            row = process_pdf_page(
                pdf_path,
                source_name,
                page_index,
                page_count,
                page_text,
                index_rows,
                fnsku_pattern,
                output_dir,
                batch_fnskus,
                feishu_token,
                feishu_config,
                remote_by_fnsku,
                delete_existing_same_fnsku,
                stage_callback=page_stage,
                defer_upload=defer_upload,
            )
            log_rows.append(row)
            if page_finished_callback:
                page_finished_callback(source_name, page_index, page_count, row)
        return log_rows

    except Exception as exc:
        row = fail_row(source_name, "处理失败", "", f"处理失败：{exc}")
        if page_finished_callback:
            page_finished_callback(source_name, 0, 1, row)
        return [row]


def write_log_workbook(log_rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "处理日志"
    worksheet.append(LOG_HEADERS)

    for row in log_rows:
        worksheet.append([row.get(header, "") for header in LOG_HEADERS])

    worksheet.freeze_panes = "A2"
    for column_index, header in enumerate(LOG_HEADERS, start=1):
        values = [header]
        values.extend(cell_to_text(row.get(header, "")) for row in log_rows)
        width = min(max(len(value) for value in values) + 2, 80)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_zip(output_dir, log_bytes):
    zip_buffer = io.BytesIO()
    log_name = f"处理日志/处理日志_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # PDFs are already compressed; storing avoids wasting time recompressing them.
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_STORED) as archive:
        archive.writestr("本地备份标签/", b"")
        archive.writestr("处理日志/", b"")
        for pdf_path in sorted(output_dir.glob("*.pdf"), key=lambda path: natural_sort_key(path.name)):
            archive.write(pdf_path, f"本地备份标签/{pdf_path.name}")
        archive.writestr(log_name, log_bytes)

    zip_buffer.seek(0)
    return zip_buffer.getvalue()


def format_page_text(page_index, page_count):
    if page_count and page_count > 1:
        return f"第 {page_index + 1}/{page_count} 页"
    return "第 1 页"


def count_pdf_pages(saved_pdfs):
    total_pages = 0
    for _, pdf_path in saved_pdfs:
        try:
            with fitz.open(pdf_path) as document:
                total_pages += max(document.page_count, 1)
        except Exception:
            total_pages += 1
    return max(total_pages, 1)


def mark_row_not_uploaded(row, reason="用户停止任务，未上传"):
    row["匹配状态"] = SKIPPED_STATUS
    row["失败原因"] = reason
    row["飞书文件token"] = ""
    row["处理动作"] = f"跳过({reason})"
    row["_needs_upload"] = False


def upload_pending_rows(
    log_rows,
    feishu_token,
    feishu_config,
    remote_by_fnsku,
    status_callback=None,
    cancel_callback=None,
):
    pending_rows = [
        row
        for row in log_rows
        if row.get("_needs_upload") and row.get("匹配状态") == SUCCESS_STATUS
    ]
    total = len(pending_rows)
    if not pending_rows:
        return {"total": 0, "done": 0, "success": 0, "failed": 0, "cancelled": False}

    worker_count = min(UPLOAD_WORKERS, total)

    def upload_one(row):
        local_path = Path(row["_local_path"])
        last_error = None
        for attempt in range(1, 4):
            try:
                return feishu_upload_file(feishu_token, feishu_config, local_path, row["飞书正式文件名"])
            except Exception as exc:
                last_error = exc
                if attempt < 3:
                    time.sleep(attempt * 1.5)
        raise last_error

    done = 0
    success_count = 0
    failed_count = 0
    next_index = 0
    cancelled = False

    def cancel_requested():
        return bool(cancel_callback and cancel_callback())

    executor = ThreadPoolExecutor(max_workers=worker_count)
    future_to_row = {}

    def submit_next():
        nonlocal next_index
        if next_index >= total:
            return False
        row = pending_rows[next_index]
        next_index += 1
        future_to_row[executor.submit(upload_one, row)] = row
        return True

    try:
        while next_index < total and len(future_to_row) < worker_count and not cancel_requested():
            submit_next()

        if next_index == 0 and cancel_requested():
            cancelled = True

        while future_to_row:
            if cancel_requested():
                cancelled = True
                break

            done_futures, _ = wait(
                future_to_row.keys(),
                timeout=1,
                return_when=FIRST_COMPLETED,
            )
            if not done_futures:
                continue

            for future in done_futures:
                row = future_to_row.pop(future)
                done += 1
                fnsku = row.get("识别FNSKU") or ""

                try:
                    file_token = future.result()
                except Exception as exc:
                    failed_count += 1
                    row["匹配状态"] = "失败"
                    row["失败原因"] = f"飞书上传失败：{exc}"
                    row["飞书文件token"] = ""
                    row["处理动作"] = "异常(飞书上传失败，未上传)"
                    row["_needs_upload"] = False
                    status_text = "上传失败"
                    last_upload_error = f"{fnsku} 上传失败：{exc}"
                else:
                    success_count += 1
                    row["飞书文件token"] = file_token
                    row["处理动作"] = "生成修改后PDF；上传飞书"
                    row["_needs_upload"] = False
                    remote_by_fnsku[fnsku] = [
                        {"name": row["飞书正式文件名"], "token": file_token, "type": "file"}
                    ]
                    status_text = "上传成功"
                    last_upload_error = ""

                if status_callback:
                    status_callback(
                        {
                            "phase": "upload",
                            "message": (
                                f"正在并发上传到飞书：{done}/{total}，"
                                f"成功 {success_count}，失败 {failed_count}，并发 {worker_count}，"
                                f"{fnsku} {status_text}"
                            ),
                            "upload_done": done,
                            "upload_total": total,
                            "upload_success_count": success_count,
                            "upload_failed_count": failed_count,
                            "upload_workers": worker_count,
                            "last_upload_error": last_upload_error,
                        }
                    )

            if cancelled:
                break

            while next_index < total and len(future_to_row) < worker_count and not cancel_requested():
                submit_next()

        if cancel_requested():
            cancelled = True
    finally:
        executor.shutdown(wait=not cancelled, cancel_futures=True)

    if cancelled:
        stopped_rows = []
        seen_row_ids = set()
        for row in list(future_to_row.values()) + pending_rows[next_index:]:
            row_id = id(row)
            if row_id in seen_row_ids:
                continue
            seen_row_ids.add(row_id)
            stopped_rows.append(row)

        for row in stopped_rows:
            mark_row_not_uploaded(row)

    if cancelled and status_callback:
        status_callback(
            {
                "phase": "upload",
                "message": (
                    f"任务已停止：已尝试上传 {done}/{total}，"
                    f"成功 {success_count}，失败 {failed_count}，剩余未上传。"
                ),
                "upload_done": done,
                "upload_total": total,
                "upload_success_count": success_count,
                "upload_failed_count": failed_count,
                "upload_workers": worker_count,
                "cancelled": True,
            }
        )

    return {
        "total": total,
        "done": done,
        "success": success_count,
        "failed": failed_count,
        "cancelled": cancelled,
    }


def process_saved_pdfs(
    saved_pdfs,
    output_dir,
    delete_existing_same_fnsku=True,
    status_callback=None,
    cancel_callback=None,
):
    def report(message, progress=None, **fields):
        if status_callback:
            payload = {"message": message}
            if progress is not None:
                payload["progress"] = progress
            payload.update(fields)
            status_callback(payload)

    def cancel_requested():
        return bool(cancel_callback and cancel_callback())

    output_dir.mkdir(parents=True, exist_ok=True)

    report("正在读取飞书配置...", 0.02, status="处理中")
    feishu_config = load_feishu_config()
    report("正在连接飞书...", 0.04, status="处理中")
    feishu_token = get_feishu_tenant_access_token(feishu_config)
    report("正在读取飞书在线索引表...", 0.07, status="处理中")
    index_result = read_feishu_sheet_index(feishu_token, feishu_config)
    index_rows = index_result["rows"]
    report(f"已读取 {len(index_rows)} 条索引，正在检查飞书目标文件夹已有标签...", 0.12)
    remote_files = feishu_list_folder_files(feishu_token, feishu_config)
    remote_by_fnsku = build_remote_fnsku_map(remote_files)

    saved_pdfs = sorted(saved_pdfs, key=lambda item: natural_sort_key(item[0]))

    report("正在扫描 PDF 页数...", 0.15)
    total_pages = count_pdf_pages(saved_pdfs)

    fnsku_pattern = re.compile(FNSKU_REGEX, re.IGNORECASE)
    log_rows = []
    batch_fnskus = set()

    processed_pages = 0
    pending_upload_count = 0
    skipped_count = 0
    error_count = 0
    cancelled = False

    def stage_callback(source_name, page_index, page_count, message):
        report(
            f"{source_name}：{format_page_text(page_index, page_count)}，{message}",
            phase="prepare",
        )

    def page_finished_callback(source_name, page_index, page_count, row):
        nonlocal processed_pages, pending_upload_count, skipped_count, error_count
        processed_pages += 1
        if row.get("匹配状态") == SUCCESS_STATUS:
            pending_upload_count += 1
        elif row.get("匹配状态") == SKIPPED_STATUS:
            skipped_count += 1
        else:
            error_count += 1

        fnsku = row.get("识别FNSKU") or "未识别"
        status = row.get("匹配状态") or "未知"
        progress_value = 0.15 + min(processed_pages / total_pages, 1.0) * 0.35
        report(
            (
                f"{source_name}：{format_page_text(page_index, page_count)}，"
                f"FNSKU：{fnsku}，结果：{status}"
            ),
            progress_value,
            phase="prepare",
            processed_pages=processed_pages,
            total_pages=total_pages,
            pending_upload_count=pending_upload_count,
            skipped_count=skipped_count,
            error_count=error_count,
        )

    report(f"开始处理，共 {len(saved_pdfs)} 个 PDF，约 {total_pages} 张标签。", 0.15)

    for index, (source_name, pdf_path) in enumerate(saved_pdfs, start=1):
        if cancel_requested():
            cancelled = True
            report("任务已收到停止请求，剩余 PDF 不再处理。", status="正在停止", cancelled=True)
            break
        report(f"正在处理第 {index}/{len(saved_pdfs)} 个 PDF：{source_name}", phase="prepare")
        log_rows.extend(
            process_pdf_file(
                pdf_path,
                source_name,
                index_rows,
                fnsku_pattern,
                output_dir,
                batch_fnskus,
                feishu_token,
                feishu_config,
                remote_by_fnsku,
                delete_existing_same_fnsku,
                stage_callback=stage_callback,
                page_finished_callback=page_finished_callback,
                defer_upload=True,
            )
        )

    pending_rows = [
        row
        for row in log_rows
        if row.get("_needs_upload") and row.get("匹配状态") == SUCCESS_STATUS
    ]
    upload_total = len(pending_rows)

    def upload_status_callback(payload):
        upload_done = payload.get("upload_done", 0)
        upload_success_count = payload.get("upload_success_count", 0)
        upload_failed_count = payload.get("upload_failed_count", 0)
        upload_total_value = max(payload.get("upload_total", upload_total), 1)
        progress_value = 0.50 + min(upload_done / upload_total_value, 1.0) * 0.40
        report(
            payload["message"],
            progress_value,
            status="正在停止" if payload.get("cancelled") else "处理中",
            phase="upload",
            upload_done=upload_done,
            upload_total=payload.get("upload_total", upload_total),
            upload_success_count=upload_success_count,
            upload_failed_count=upload_failed_count,
            upload_workers=payload.get("upload_workers", UPLOAD_WORKERS),
            last_upload_error=payload.get("last_upload_error", ""),
            error_count=error_count + upload_failed_count,
            success_count=upload_success_count,
            cancelled=payload.get("cancelled", False),
        )

    upload_stats = {"success": 0, "failed": 0, "cancelled": cancelled}
    if cancelled:
        for row in pending_rows:
            mark_row_not_uploaded(row)
        report("任务已停止，已生成但未上传的标签会写入日志。", 0.90, status="正在停止", cancelled=True)
    elif upload_total:
        report(f"开始并发上传到飞书：共 {upload_total} 个，并发 {min(UPLOAD_WORKERS, upload_total)}。", 0.50)
        upload_stats = upload_pending_rows(
            log_rows,
            feishu_token,
            feishu_config,
            remote_by_fnsku,
            upload_status_callback,
            cancel_callback=cancel_callback,
        )
        cancelled = cancelled or upload_stats.get("cancelled", False)

    if cancelled:
        report("任务已停止，正在生成处理日志...", 0.96, status="正在停止", cancelled=True)
        cleanup_deleted = []
    else:
        report("处理完成，正在检查飞书文件夹是否还有重复 FNSKU...", 0.92)
        cleanup_deleted = cleanup_remote_duplicates(feishu_token, remote_by_fnsku)
    report("正在生成处理日志...", 0.96)
    log_bytes = write_log_workbook(log_rows)
    report("正在生成本次 PDF 备份 ZIP...", 0.98)
    zip_bytes = build_zip(output_dir, log_bytes)

    summary = summarize_result({"log_rows": log_rows, "cleanup_deleted": cleanup_deleted})
    cleanup_text = f"，清理重复 {len(cleanup_deleted)} 个" if cleanup_deleted else ""
    final_status = "已停止" if cancelled else "完成"
    final_text = (
        f"已停止：已上传 {summary['success_count']} 张，跳过 {summary['skipped_count']} 张，"
        f"异常 {summary['error_count']} 张。"
        if cancelled
        else (
            f"全部完成：上传 {summary['success_count']} 张，跳过 {summary['skipped_count']} 张，"
            f"异常 {summary['error_count']} 张{cleanup_text}。"
        )
    )
    report(
        final_text,
        1.0,
        status=final_status,
        cancelled=cancelled,
        **summary,
    )

    return {
        "warnings": index_result["warnings"],
        "used_sheets": index_result["used_sheets"],
        "skipped_sheets": index_result["skipped_sheets"],
        "remote_fnsku_count": len(remote_by_fnsku),
        "cleanup_deleted": cleanup_deleted,
        "log_rows": log_rows,
        "log_bytes": log_bytes,
        "zip_bytes": zip_bytes,
        "cancelled": cancelled,
    }


def save_uploaded_files_to_dir(pdf_files, input_dir):
    input_dir.mkdir(parents=True, exist_ok=True)
    saved_pdfs = []
    for uploaded_pdf in pdf_files:
        filename = safe_upload_name(uploaded_pdf.name, ".pdf")
        if not filename.lower().endswith(".pdf"):
            filename = f"{filename}.pdf"
        pdf_path = unique_path(input_dir, filename)
        pdf_path.write_bytes(uploaded_pdf.getvalue())
        saved_pdfs.append((uploaded_pdf.name, pdf_path))
    return sorted(saved_pdfs, key=lambda item: natural_sort_key(item[0]))


def process_uploads(pdf_files, delete_existing_same_fnsku=True):
    status_box = st.empty()
    progress = st.progress(0, text="准备开始...")

    def status_callback(payload):
        message = payload.get("message", "")
        progress_value = payload.get("progress")
        if progress_value is not None:
            progress.progress(
                min(max(float(progress_value), 0.0), 1.0),
                text=message,
            )
        status_box.info(message)

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_root = Path(temp_dir)
        input_dir = temp_root / "input"
        output_dir = temp_root / "output"
        status_box.info("正在保存上传的 PDF...")
        saved_pdfs = save_uploaded_files_to_dir(pdf_files, input_dir)
        result = process_saved_pdfs(
            saved_pdfs,
            output_dir,
            delete_existing_same_fnsku=delete_existing_same_fnsku,
            status_callback=status_callback,
        )
        status_box.success("全部完成。")
        return result


def summarize_result(result):
    log_rows = result.get("log_rows", [])
    success_count = sum(1 for row in log_rows if row.get("匹配状态") == SUCCESS_STATUS)
    skipped_count = sum(1 for row in log_rows if row.get("匹配状态") == SKIPPED_STATUS)
    error_count = sum(
        1
        for row in log_rows
        if row.get("匹配状态") not in (SUCCESS_STATUS, SKIPPED_STATUS)
    )
    return {
        "success_count": success_count,
        "skipped_count": skipped_count,
        "error_count": error_count,
        "total_count": len(log_rows),
        "cleanup_deleted_count": len(result.get("cleanup_deleted", [])),
    }


def prune_process_history():
    if not PROCESS_HISTORY_DIR.exists():
        return

    run_dirs = [path for path in PROCESS_HISTORY_DIR.iterdir() if path.is_dir()]
    run_dirs.sort(key=lambda path: path.name, reverse=True)
    for old_dir in run_dirs[PROCESS_HISTORY_LIMIT:]:
        shutil.rmtree(old_dir, ignore_errors=True)


def save_process_history(result):
    PROCESS_HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:8]
    run_dir = PROCESS_HISTORY_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    summary = summarize_result(result)
    metadata = {
        "run_id": run_id,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "used_sheets": result.get("used_sheets", []),
        "skipped_sheets": result.get("skipped_sheets", []),
        "warnings": result.get("warnings", []),
        "cleanup_deleted": result.get("cleanup_deleted", [])[:20],
        **summary,
    }

    (run_dir / "metadata.json").write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (run_dir / "处理日志.xlsx").write_bytes(result.get("log_bytes", b""))
    (run_dir / "本次生成PDF备份.zip").write_bytes(result.get("zip_bytes", b""))
    prune_process_history()
    return metadata


def read_json_file(path, default=None):
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def write_json_file(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    temp_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temp_path, path)


def job_metadata_path(job_dir):
    return job_dir / "metadata.json"


def job_cancel_path(job_dir):
    return job_dir / "cancel.requested"


def update_job_metadata(job_dir, **updates):
    metadata_path = job_metadata_path(job_dir)
    metadata = read_json_file(metadata_path, {}) or {}
    if metadata.get("cancel_requested") and updates.get("status") == "处理中":
        updates["status"] = "正在停止"
    metadata.update(updates)
    metadata["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    write_json_file(metadata_path, metadata)
    return metadata


def is_job_cancel_requested(job_dir):
    metadata = read_json_file(job_metadata_path(job_dir), {}) or {}
    return bool(metadata.get("cancel_requested") or job_cancel_path(job_dir).exists())


def request_job_cancel(job_dir):
    job_cancel_path(job_dir).write_text(datetime.now().isoformat(), encoding="utf-8")
    return update_job_metadata(
        job_dir,
        cancel_requested=True,
        status="正在停止",
        message="已请求停止，正在结束任务，不会继续上传后续文件。",
    )


def prune_job_history():
    if not JOB_DIR.exists():
        return

    job_dirs = [path for path in JOB_DIR.iterdir() if path.is_dir()]
    job_dirs.sort(key=lambda path: path.name, reverse=True)
    for old_dir in job_dirs[JOB_HISTORY_LIMIT:]:
        shutil.rmtree(old_dir, ignore_errors=True)


def create_processing_job(pdf_files):
    JOB_DIR.mkdir(parents=True, exist_ok=True)
    job_id = datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:8]
    job_dir = JOB_DIR / job_id
    input_dir = job_dir / "input"
    output_dir = job_dir / "output"
    input_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    saved_pdfs = save_uploaded_files_to_dir(pdf_files, input_dir)
    file_records = [{"source_name": source_name, "path": str(path)} for source_name, path in saved_pdfs]
    write_json_file(job_dir / "files.json", file_records)

    metadata = {
        "job_id": job_id,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "status": "排队中",
        "message": "任务已提交，后台准备开始处理。",
        "progress": 0.0,
        "pdf_count": len(saved_pdfs),
        "upload_workers": UPLOAD_WORKERS,
        "cancel_requested": False,
    }
    write_json_file(job_metadata_path(job_dir), metadata)
    prune_job_history()

    thread = threading.Thread(target=run_processing_job, args=(job_id,), daemon=True)
    thread.start()
    return metadata


def run_processing_job(job_id):
    job_dir = JOB_DIR / job_id
    output_dir = job_dir / "output"
    file_records = read_json_file(job_dir / "files.json", []) or []
    saved_pdfs = [(record["source_name"], Path(record["path"])) for record in file_records]

    def status_callback(payload):
        allowed_keys = {
            "status",
            "message",
            "progress",
            "phase",
            "processed_pages",
            "total_pages",
            "pending_upload_count",
            "skipped_count",
            "error_count",
            "upload_done",
            "upload_total",
            "upload_success_count",
            "upload_failed_count",
            "upload_workers",
            "success_count",
            "total_count",
            "cleanup_deleted_count",
            "last_upload_error",
            "cancelled",
        }
        clean_payload = {key: payload[key] for key in allowed_keys if key in payload}
        if is_job_cancel_requested(job_dir) and clean_payload.get("status") == "处理中":
            clean_payload["status"] = "正在停止"
        update_job_metadata(job_dir, **clean_payload)

    try:
        update_job_metadata(job_dir, status="处理中", message="后台任务已开始。", progress=0.01)
        result = process_saved_pdfs(
            saved_pdfs,
            output_dir,
            status_callback=status_callback,
            cancel_callback=lambda: is_job_cancel_requested(job_dir),
        )

        (job_dir / "处理日志.xlsx").write_bytes(result["log_bytes"])
        (job_dir / "本次生成PDF备份.zip").write_bytes(result["zip_bytes"])
        try:
            save_process_history(result)
        except Exception:
            pass

        summary = summarize_result(result)
        if result.get("cancelled"):
            update_job_metadata(
                job_dir,
                status="已停止",
                message=(
                    f"已停止：上传 {summary['success_count']}，跳过 {summary['skipped_count']}，"
                    f"异常 {summary['error_count']}。"
                ),
                progress=1.0,
                cancelled=True,
                **summary,
            )
        else:
            update_job_metadata(
                job_dir,
                status="完成",
                message=(
                    f"完成：上传 {summary['success_count']}，跳过 {summary['skipped_count']}，"
                    f"异常 {summary['error_count']}。"
                ),
                progress=1.0,
                **summary,
            )
    except Exception as exc:
        update_job_metadata(
            job_dir,
            status="失败",
            message=f"任务失败：{exc}",
            error=str(exc),
            progress=1.0,
        )


def load_job_history():
    if not JOB_DIR.exists():
        return []

    jobs = []
    for job_dir in sorted([path for path in JOB_DIR.iterdir() if path.is_dir()], key=lambda p: p.name, reverse=True):
        metadata = read_json_file(job_metadata_path(job_dir), None)
        if not metadata:
            continue
        metadata["_job_dir"] = str(job_dir)
        jobs.append(metadata)
    return jobs[:JOB_HISTORY_LIMIT]


def render_jobs():
    jobs = load_job_history()
    if not jobs:
        return

    st.divider()
    cols = st.columns([1, 4])
    with cols[0]:
        st.button("刷新任务状态", use_container_width=True)
    with cols[1]:
        st.subheader("后台任务")
        st.caption("页面不会自动刷新。需要看最新进度时，点左侧刷新按钮即可。")

    for index, job in enumerate(jobs, start=1):
        title = (
            f"{job.get('created_at', '未知时间')}｜{job.get('status', '未知')}｜"
            f"{job.get('message', '')}"
        )
        with st.expander(title, expanded=index == 1):
            job_dir = Path(job.get("_job_dir", ""))
            is_active = job.get("status") in {"排队中", "处理中", "正在停止"}
            progress_value = float(job.get("progress") or 0)
            st.progress(min(max(progress_value, 0.0), 1.0), text=job.get("message", ""))
            cols = st.columns(6)
            cols[0].metric("PDF", job.get("pdf_count", 0))
            cols[1].metric("成功上传", job.get("success_count", job.get("upload_success_count", 0)))
            cols[2].metric("已尝试", job.get("upload_done", 0))
            cols[3].metric("跳过", job.get("skipped_count", 0))
            cols[4].metric("异常", job.get("error_count", job.get("upload_failed_count", 0)))
            cols[5].metric("并发", job.get("upload_workers", UPLOAD_WORKERS))

            if is_active and not job.get("cancel_requested"):
                if st.button(
                    "停止这个任务",
                    type="secondary",
                    use_container_width=True,
                    key=f"cancel-job-{job.get('job_id', index)}",
                ):
                    request_job_cancel(job_dir)
                    st.warning("已发出停止请求。当前正在上传的几个文件会先收尾，后面的不会继续上传。")
                    st.rerun()
            elif job.get("cancel_requested") and is_active:
                st.warning("正在停止：等待当前上传请求收尾。")

            if job.get("status") == "失败":
                st.error(job.get("error") or job.get("message"))
            if job.get("last_upload_error"):
                st.warning("最近一次上传失败：" + job.get("last_upload_error"))

            log_path = job_dir / "处理日志.xlsx"
            zip_path = job_dir / "本次生成PDF备份.zip"
            if log_path.exists():
                st.download_button(
                    "下载处理日志 Excel",
                    data=log_path.read_bytes(),
                    file_name=f"处理日志_{job.get('job_id', index)}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"job-log-{job.get('job_id', index)}",
                )
            if zip_path.exists():
                st.download_button(
                    "下载 PDF 备份 ZIP",
                    data=zip_path.read_bytes(),
                    file_name=f"标签本地备份_{job.get('job_id', index)}.zip",
                    mime="application/zip",
                    use_container_width=True,
                    key=f"job-zip-{job.get('job_id', index)}",
                )


def load_process_history():
    if not PROCESS_HISTORY_DIR.exists():
        return []

    history = []
    run_dirs = [path for path in PROCESS_HISTORY_DIR.iterdir() if path.is_dir()]
    run_dirs.sort(key=lambda path: path.name, reverse=True)
    for run_dir in run_dirs[:PROCESS_HISTORY_LIMIT]:
        metadata_path = run_dir / "metadata.json"
        if not metadata_path.exists():
            continue
        try:
            metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        except Exception:
            continue

        metadata["_run_dir"] = str(run_dir)
        history.append(metadata)
    return history


def render_history():
    history = load_process_history()
    if not history:
        return

    st.divider()
    st.subheader("最近三次处理记录")
    for index, item in enumerate(history, start=1):
        title = (
            f"{item.get('created_at', '未知时间')}："
            f"上传 {item.get('success_count', 0)}，"
            f"跳过 {item.get('skipped_count', 0)}，"
            f"异常 {item.get('error_count', 0)}"
        )
        with st.expander(title, expanded=index == 1):
            st.caption("刷新页面后也可以从这里重新下载最近的处理结果。")
            cols = st.columns(4)
            cols[0].metric("上传成功", item.get("success_count", 0))
            cols[1].metric("已跳过", item.get("skipped_count", 0))
            cols[2].metric("异常", item.get("error_count", 0))
            cols[3].metric("清理重复", item.get("cleanup_deleted_count", 0))

            run_dir = Path(item.get("_run_dir", ""))
            log_path = run_dir / "处理日志.xlsx"
            zip_path = run_dir / "本次生成PDF备份.zip"
            if log_path.exists():
                st.download_button(
                    "下载这次处理日志 Excel",
                    data=log_path.read_bytes(),
                    file_name=f"处理日志_{item.get('run_id', index)}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"history-log-{item.get('run_id', index)}",
                )
            if zip_path.exists():
                st.download_button(
                    "下载这次 PDF 备份 ZIP",
                    data=zip_path.read_bytes(),
                    file_name=f"标签本地备份_{item.get('run_id', index)}.zip",
                    mime="application/zip",
                    use_container_width=True,
                    key=f"history-zip-{item.get('run_id', index)}",
                )


def render_result(result):
    log_rows = result["log_rows"]
    success_count = sum(1 for row in log_rows if row["匹配状态"] == SUCCESS_STATUS)
    skipped_rows = [row for row in log_rows if row["匹配状态"] == SKIPPED_STATUS]
    error_rows = [
        row for row in log_rows if row["匹配状态"] not in (SUCCESS_STATUS, SKIPPED_STATUS)
    ]

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("成功上传", success_count)
    col2.metric("已跳过", len(skipped_rows))
    col3.metric("异常标签", len(error_rows))
    col4.metric("总标签", len(log_rows))

    if result["used_sheets"]:
        st.info("已读取索引 Sheet：" + "、".join(result["used_sheets"]))
    if result["skipped_sheets"]:
        st.caption("已跳过非索引 Sheet：" + "、".join(result["skipped_sheets"]))
    for warning in result["warnings"]:
        st.warning(warning)
    cleanup_deleted = result.get("cleanup_deleted", [])
    if cleanup_deleted:
        st.info(f"飞书重复检查：已清理 {len(cleanup_deleted)} 个重复文件。")
    else:
        st.caption("飞书重复检查：未发现需要清理的重复文件。")

    if error_rows:
        st.error("有异常标签，异常标签没有上传成功，请查看日志。")
        st.dataframe(
            [
                {
                    "原文件名": row["原文件名"],
                    "页码": row["页码"],
                    "识别FNSKU": row["识别FNSKU"],
                    "失败原因": row["失败原因"],
                }
                for row in error_rows
            ],
            use_container_width=True,
            hide_index=True,
        )
    else:
        st.success("处理完成，没有异常标签。")

    reason_counter = Counter(row["失败原因"] for row in error_rows)
    if reason_counter:
        st.caption("异常统计：" + "；".join(f"{reason}：{count}" for reason, count in reason_counter.items()))

    st.download_button(
        "下载处理日志 Excel",
        data=result["log_bytes"],
        file_name=f"处理日志_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.download_button(
        "下载本次生成 PDF 备份 ZIP",
        data=result["zip_bytes"],
        file_name=f"标签本地备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
        mime="application/zip",
        use_container_width=True,
    )


def main():
    st.set_page_config(page_title="亚马逊标签 PDF 飞书上传工具", layout="wide")
    st.title("亚马逊标签 PDF 飞书上传工具")

    pdf_files = st.file_uploader(
        "上传标签 PDF（可多选，也支持一个文件里有多页标签）",
        type=["pdf"],
        accept_multiple_files=True,
    )

    can_process = bool(pdf_files)
    if st.button("提交后台处理任务", type="primary", disabled=not can_process, use_container_width=True):
        try:
            with st.spinner("正在提交任务，请稍等..."):
                job = create_processing_job(pdf_files)
                st.session_state["last_job_id"] = job["job_id"]
            st.success(f"任务已提交：{job['job_id']}。可以关闭页面，之后回来刷新查看结果。")
        except FatalError as exc:
            st.error(f"启动失败：{exc}")
        except FeishuError as exc:
            st.error(f"飞书接口失败：{exc}")
        except Exception as exc:
            st.error(f"程序异常停止：{exc}")

    render_jobs()
    render_history()


if __name__ == "__main__":
    main()
